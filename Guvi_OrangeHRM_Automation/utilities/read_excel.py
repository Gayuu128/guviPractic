from openpyxl import load_workbook


class ReadExcel:

    @staticmethod
    def get_row_count(file, sheet_name):
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def read_data(file, sheet_name, row, column):
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row, column=column).value

    @staticmethod
    def write_data(file, sheet_name, row, column, data):
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=column).value = data
        workbook.save(file)