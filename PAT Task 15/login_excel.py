from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime


# =======================
# Excel Utility
# =======================
class ExcelUtils:

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook.active

    def get_data(self, row, col):
        return self.sheet.cell(row=row, column=col).value

    def write_result(self, row, result):
        self.sheet.cell(row=row, column=4).value = str(datetime.now().date())
        self.sheet.cell(row=row, column=5).value = datetime.now().strftime("%H:%M:%S")
        self.sheet.cell(row=row, column=7).value = result
        self.workbook.save(self.file_path)


# =======================
# Login Page (POM)
# =======================
class LoginPage:

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 5)

    username_loc = (By.NAME, "username")
    password_loc = (By.NAME, "password")
    login_btn_loc = (By.XPATH, "//button[@type='submit']")
    dashboard_loc = (By.XPATH, "//h6[text()='Dashboard']")
    error_loc = (By.XPATH, "//p[contains(@class,'oxd-alert-content-text')]")

    def login(self, username, password):
        self.wait.until(EC.visibility_of_element_located(self.username_loc)).send_keys(username)
        self.wait.until(EC.visibility_of_element_located(self.password_loc)).send_keys(password)
        self.wait.until(EC.element_to_be_clickable(self.login_btn_loc)).click()

    def is_login_success(self):
        try:
            self.wait.until(EC.visibility_of_element_located(self.dashboard_loc))
            return True
        except:
            return False